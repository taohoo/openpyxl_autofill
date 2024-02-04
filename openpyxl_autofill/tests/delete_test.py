# -*- coding: utf-8 -*-
"""
@author: hubo
@project: openpyxl_autofill
@file: insert_test.py
@time: 2023/12/22 14:15
@desc:
"""
from openpyxl import load_workbook
from .. import enable_all
enable_all()


def test_delete_cols():
    wb = load_workbook('openpyxl_autofill/tests/test.xlsx')
    ws = wb.worksheets[0]
    ws.delete_cols(2, amount=2)
    wb.save('openpyxl_autofill/tests/out.xlsx')


def test_delete_rows():
    wb = load_workbook('openpyxl_autofill/tests/test.xlsx')
    ws = wb.worksheets[0]
    ws.delete_rows(4)
    wb.save('openpyxl_autofill/tests/out.xlsx')



