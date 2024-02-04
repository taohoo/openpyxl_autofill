# -*- coding: utf-8 -*-
"""
@author: hubo
@project: openpyxl_autofill
@file: insert_test.py
@time: 2023/12/22 14:15
@desc:
"""
from openpyxl import Workbook, load_workbook
from .. import enable_all
enable_all()


def test_insert_cols():
    wb = load_workbook('openpyxl_autofill/tests/test.xlsx')
    ws = wb.worksheets[0]
    # ws.insert_rows(6)
    ws.insert_cols(2, amount=2)
    wb.save('openpyxl_autofill/tests/out.xlsx')


def test_insert_rows():
    wb = load_workbook('openpyxl_autofill/tests/test.xlsx')
    ws = wb.worksheets[0]
    # ws.insert_rows(6)
    # ws.insert_rows(4)
    ws.insert_rows(4)
    wb.save('openpyxl_autofill/tests/out.xlsx')



