# -*- coding: utf-8 -*-
"""
@author: hubo
@project: bb-py
@file: insert_test.py
@time: 2023/12/22 14:15
@desc:
"""
from openpyxl import load_workbook
from .. import patch_all
patch_all()


def test_insert_cols():
    wb = load_workbook('bb_openpyxl/tests/test.xlsx')
    ws = wb.worksheets[0]
    # ws.insert_rows(6)
    ws.insert_cols(2, amount=2)
    wb.save('bb_openpyxl/tests/out.xlsx')


def test_insert_rows():
    wb = load_workbook('bb_openpyxl/tests/test.xlsx')
    ws = wb.worksheets[0]
    # ws.insert_rows(6)
    # ws.insert_rows(4)
    ws.insert_rows(4)
    wb.save('bb_openpyxl/tests/out.xlsx')



