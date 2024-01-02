# -*- coding: utf-8 -*-
"""
@author: hubo
@project: bb-py
@file: _formula.py
@time: 2023/12/23 11:35
@desc:
"""
import re
from openpyxl.utils import column_index_from_string, get_column_letter
from ._utils import _duplicate


def _get_all_cells(formula):
    """提取公式中的单元格"""
    cell_pattern = "[^A-Z0-9]([A-Z]+[0-9]+)"
    return _duplicate(re.findall(cell_pattern, formula))


def _get_all_rows(formula):
    """提取公式中的行"""
    row_pattern = r"\$(\d+)"
    return _duplicate(re.findall(row_pattern, formula))


def _get_all_columns(formula):
    """提取公式中的列"""
    column_pattern = r"\$([A-Z]+)"
    return _duplicate(re.findall(column_pattern, formula))


def _calculate_new_cell(cell, row_idx=None, col_idx=None, amount=1):
    col = column_index_from_string(re.findall(r"([A-Z]+)", cell)[0])
    row = int(re.findall(r"(\d+)", cell)[0])
    if row_idx is not None and row_idx <= row:
        row += amount
    if col_idx is not None and col_idx <= col:
        col += amount
    return '%s%d' % (get_column_letter(col), row)


def _reset_cells_in_formula(formula, cells, current_cell=None, row_idx=None, col_idx=None, amount=1):
    new_formula = formula
    if current_cell is None and len(cells) > 0:
        current_cell = cells.pop(0)
    if not current_cell:
        return new_formula
    new_cell = _calculate_new_cell(current_cell, row_idx=row_idx, col_idx=col_idx, amount=amount)
    if new_cell != current_cell:
        if new_cell in cells:
            # 要把new_cell处理掉
            cells.remove(new_cell)
            new_formula = _reset_cells_in_formula(new_formula, cells, new_cell, row_idx=row_idx, col_idx=col_idx, amount=amount)
        new_formula = new_formula.replace(current_cell, new_cell)
    return _reset_cells_in_formula(new_formula, cells, row_idx=row_idx, col_idx=col_idx, amount=amount)


def _calculate_new_col(col, idx, amount=1):
    current = column_index_from_string(col)
    if idx <= current:
        return get_column_letter(current + amount)
    return col

def _reset_cols_in_formula(formula, cols, current_col=None, idx=None, amount=1):
    new_formula = formula
    if current_col is None and len(cols) > 0:
        current_col = cols.pop(0)
    if not current_col:
        return new_formula
    new_col = _calculate_new_col(current_col, idx, amount=amount)
    if new_col != current_col:
        if new_col in cols:
            # 要把new_cell处理掉
            cols.remove(new_col)
            new_formula = _reset_cols_in_formula(new_formula, cols, new_col, idx=idx, amount=amount)
        new_formula = new_formula.replace('$'+current_col, '$'+new_col)
    return _reset_cols_in_formula(new_formula, cols, idx=idx, amount=amount)


def _calculate_new_row(row, idx, amount=1):
    current = int(row)
    if idx <= current:
        return str(current + amount)
    return row


def _reset_rows_in_formula(formula, rows, current_row=None, idx=None, amount=1):
    new_formula = formula
    if current_row is None and len(rows) > 0:
        current_row = rows.pop(0)
    if not current_row:
        return new_formula
    new_row = _calculate_new_row(current_row, idx, amount=amount)
    if new_row != current_row:
        if new_row in rows:
            # 要把new_cell处理掉
            rows.remove(new_row)
            new_formula = _reset_rows_in_formula(new_formula, rows, new_row, idx=idx, amount=amount)
        new_formula = new_formula.replace('$'+current_row, '$'+new_row)
    return _reset_rows_in_formula(new_formula, rows, idx=idx, amount=amount)


def _reset_formula(formula, row_idx=None, col_idx=None, amount=1):
    new_formula = formula
    # 处理单元格
    cells_in_formula = _get_all_cells(new_formula)
    new_formula = _reset_cells_in_formula(new_formula, cells_in_formula, row_idx=row_idx, col_idx=col_idx, amount=amount)
    # 处理单列
    if col_idx is not None:
        cols_in_formala = _get_all_columns(new_formula)
        new_formula = _reset_cols_in_formula(new_formula, cols_in_formala, idx=col_idx, amount=amount)
    # 处理单行
    if row_idx is not None:
        rows_in_formala = _get_all_rows(new_formula)
        new_formula = _reset_rows_in_formula(new_formula, rows_in_formala, idx=row_idx, amount=amount)

    return new_formula
